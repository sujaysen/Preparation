import openpyxl
import json
from datetime import date
today = date.today()
date = input("Enter date : ")
if(date):
  path = date+".xlsx"
else:
  path = str(today)+".xlsx"
print(path)
path1 = "ind_nifty100list.xlsx"
wb_obj = openpyxl.load_workbook(path)
wb_obj1 = openpyxl.load_workbook(path1)
sheet_obj = wb_obj.active
sheet_obj1 = wb_obj1.active
rows = sheet_obj.max_row
columns = sheet_obj.max_column
res = []
for sym in range(sheet_obj1.max_row):
  given_symbol = sheet_obj1.cell(sym+1, 3).value
  for row in range(rows):
    if (sheet_obj.cell(row+1, 2).value == given_symbol):
      d = dict()
      for col in range(columns):
        cell_obj = sheet_obj.cell(row+1, col+1)
        d[sheet_obj.cell(1,col+1).value] = cell_obj.value
      res.append(d)

max_vol = -2000
for entry in res:
  print(entry['Symbol'],"------->>>>>>",entry['Underlying Annualised Volatility (F) = E*Sqrt(365)'])
